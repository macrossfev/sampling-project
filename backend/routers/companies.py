from typing import Optional, List
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from backend.database import get_db
from backend.models import Company, Contract
from backend.schemas import CompanyCreate, CompanyUpdate, CompanyResponse

router = APIRouter(prefix="/api/companies", tags=["companies"])


@router.get("", response_model=List[CompanyResponse])
def list_companies(
    search: Optional[str] = Query(None, description="Search by company name"),
    db: Session = Depends(get_db),
):
    query = db.query(Company)
    if search:
        query = query.filter(Company.name.contains(search))
    return query.order_by(Company.id).all()


@router.post("", response_model=CompanyResponse, status_code=201)
def create_company(data: CompanyCreate, db: Session = Depends(get_db)):
    company = Company(**data.model_dump())
    db.add(company)
    db.commit()
    db.refresh(company)
    return company


@router.get("/{company_id}", response_model=CompanyResponse)
def get_company(company_id: int, db: Session = Depends(get_db)):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    return company


@router.put("/{company_id}", response_model=CompanyResponse)
def update_company(
    company_id: int, data: CompanyUpdate, db: Session = Depends(get_db)
):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(company, key, value)
    db.commit()
    db.refresh(company)
    return company


@router.delete("/{company_id}")
def delete_company(company_id: int, db: Session = Depends(get_db)):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    # Check for linked contracts
    contract_count = (
        db.query(Contract).filter(Contract.company_id == company_id).count()
    )
    if contract_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete company: {contract_count} contracts linked. Delete contracts first.",
        )
    db.delete(company)
    db.commit()
    return {"detail": "Company deleted"}


@router.get("/import/template")
def company_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "公司导入"
    headers = ["公司名称", "所属集团", "地址", "联系人", "联系电话"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="示例水务公司")
    ws.cell(row=2, column=2, value="示例集团")
    for i, w in enumerate([20, 20, 30, 12, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=company_template.xlsx"},
    )


@router.post("/import/excel")
def import_companies(file: UploadFile = File(...), db: Session = Depends(get_db)):
    wb = load_workbook(BytesIO(file.file.read()))
    ws = wb.active
    created = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        if not name:
            continue
        exists = db.query(Company).filter(Company.name == name).first()
        if exists:
            skipped += 1
            continue
        company = Company(
            name=name,
            group_name=str(row[1]).strip() if len(row) > 1 and row[1] else None,
            address=str(row[2]).strip() if len(row) > 2 and row[2] else None,
            contact_person=str(row[3]).strip() if len(row) > 3 and row[3] else None,
            contact_phone=str(row[4]).strip() if len(row) > 4 and row[4] else None,
        )
        db.add(company)
        created += 1
    db.commit()
    return {"detail": f"导入完成：新增 {created} 家，跳过 {skipped} 家（已存在）"}
