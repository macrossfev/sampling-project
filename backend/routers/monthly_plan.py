from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session

from backend.database import get_db
from backend.models import SamplingTrip, Company
from backend.schemas import (
    SamplingTripCreate,
    SamplingTripUpdate,
    SamplingTripResponse,
    MonthlyPlanGenerate,
)
from backend.services.monthly_planner import generate_monthly_plan

router = APIRouter(prefix="/api/monthly-plan", tags=["monthly-plan"])


@router.get("/trips", response_model=List[SamplingTripResponse])
def list_trips(
    year: int = Query(...),
    month: int = Query(...),
    group_no: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(SamplingTrip).filter(
        SamplingTrip.year == year,
        SamplingTrip.month == month,
    )
    if group_no is not None:
        query = query.filter(SamplingTrip.group_no == group_no)
    trips = query.order_by(SamplingTrip.start_date, SamplingTrip.group_no).all()
    results = []
    for t in trips:
        data = SamplingTripResponse.model_validate(t)
        company = db.query(Company).filter(Company.id == t.company_id).first()
        data.company_name = company.name if company else None
        data.company_short_name = company.short_name if company else None
        results.append(data)
    return results


@router.post("/generate")
def generate_plan(data: MonthlyPlanGenerate, db: Session = Depends(get_db)):
    if data.scheme not in ("compact", "balanced", "relaxed"):
        raise HTTPException(status_code=400, detail="scheme must be compact/balanced/relaxed")
    result = generate_monthly_plan(db, data.year, data.month, data.scheme, data.group_first)
    return result


@router.get("/export-excel")
def export_excel(
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
):
    import calendar as _cal

    trips = db.query(SamplingTrip).filter(
        SamplingTrip.year == year,
        SamplingTrip.month == month,
    ).order_by(SamplingTrip.start_date, SamplingTrip.group_no).all()

    max_group = max((t.group_no for t in trips), default=1)
    group_names = ['第一组','第二组','第三组','第四组','第五组','第六组','第七组','第八组']

    # Build day map: every day of month, with two-day trips split into 去程/返程
    _, days_in_month = _cal.monthrange(year, month)
    from datetime import date as _date
    day_map = {}
    for d in range(1, days_in_month + 1):
        day_map[_date(year, month, d)] = {}

    for t in trips:
        company = db.query(Company).filter(Company.id == t.company_id).first()
        cname = (company.short_name or company.name) if company else '--'
        route_lines = (t.route_notes or '').split('\n')

        if t.trip_type == 'two_day' and t.start_date != t.end_date:
            # Outbound day
            if t.start_date in day_map:
                day_map[t.start_date][t.group_no] = {
                    'company': cname, 'trip_type': '去程',
                    'route': route_lines[0] if route_lines else '',
                    'samples': t.sampling_notes or '',
                }
            # Return day
            if t.end_date in day_map:
                day_map[t.end_date][t.group_no] = {
                    'company': cname, 'trip_type': '返程',
                    'route': route_lines[1] if len(route_lines) > 1 else '',
                    'samples': '',
                }
        else:
            if t.start_date in day_map:
                day_map[t.start_date][t.group_no] = {
                    'company': cname, 'trip_type': '当日',
                    'route': t.route_notes or '',
                    'samples': t.sampling_notes or '',
                }

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月采样计划"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    weekday_names = ['一','二','三','四','五','六','日']

    # Header row
    headers = ['日期'] + group_names[:max_group]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ws.column_dimensions['A'].width = 12
    for i in range(max_group):
        ws.column_dimensions[chr(66+i)].width = 30

    weekend_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    for row_idx, d in enumerate(range(1, days_in_month + 1), 2):
        dt = _date(year, month, d)
        wd = dt.weekday()
        is_weekend = wd >= 5
        label = f"{month}.{d} {weekday_names[wd]}"
        has_trips = bool(day_map.get(dt))

        cell = ws.cell(row=row_idx, column=1, value=label)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=not is_weekend, color="BBBBBB" if is_weekend else "000000")
        cell.border = thin_border
        if is_weekend:
            cell.fill = weekend_fill

        if has_trips:
            ws.row_dimensions[row_idx].height = 60
        else:
            ws.row_dimensions[row_idx].height = 18

        for g in range(1, max_group + 1):
            info = day_map.get(dt, {}).get(g)
            if info:
                text = f"[{info['trip_type']}] {info['company']}\n{info['route']}\n{info['samples']}"
                cell = ws.cell(row=row_idx, column=g+1, value=text.strip())
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell = ws.cell(row=row_idx, column=g+1, value='')
            cell.border = thin_border
            if is_weekend:
                cell.fill = weekend_fill

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"sampling_plan_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/trips", response_model=SamplingTripResponse)
def create_trip(data: SamplingTripCreate, db: Session = Depends(get_db)):
    trip = SamplingTrip(**data.model_dump())
    db.add(trip)
    db.commit()
    db.refresh(trip)
    company = db.query(Company).filter(Company.id == trip.company_id).first()
    resp = SamplingTripResponse.model_validate(trip)
    resp.company_name = company.name if company else None
    resp.company_short_name = company.short_name if company else None
    return resp


@router.put("/trips/{trip_id}", response_model=SamplingTripResponse)
def update_trip(trip_id: int, data: SamplingTripUpdate, db: Session = Depends(get_db)):
    trip = db.query(SamplingTrip).filter(SamplingTrip.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(trip, key, value)
    db.commit()
    db.refresh(trip)
    company = db.query(Company).filter(Company.id == trip.company_id).first()
    resp = SamplingTripResponse.model_validate(trip)
    resp.company_name = company.name if company else None
    resp.company_short_name = company.short_name if company else None
    return resp


@router.delete("/trips/{trip_id}")
def delete_trip(trip_id: int, db: Session = Depends(get_db)):
    trip = db.query(SamplingTrip).filter(SamplingTrip.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    db.delete(trip)
    db.commit()
    return {"detail": "Trip deleted"}
