import re
from io import BytesIO
from datetime import date
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from backend.models import Contract, WaterPlant, DetectionItem
from backend.services.plan_generator import generate_annual_plan


# Expected header columns (Chinese)
EXPECTED_HEADERS = [
    "水厂名称",
    "水厂规模",
    "样品类别",
    "检测项目",
    "检测标准",
    "频率类型",
    "频率值",
    "单价",
    "年检测次数",
    "小计",
]


def parse_excel(file_bytes: bytes) -> dict:
    """Parse an uploaded Excel file into a structured dict.

    Returns:
        {
            "water_plants": [
                {
                    "name": str,
                    "scale": str,
                    "items": [
                        {
                            "sample_type": str,
                            "detection_project": str,
                            "detection_standard": str,
                            "frequency_type": str,
                            "frequency_value": int,
                            "unit_price": float,
                            "annual_count": int,
                            "subtotal": float,
                        }
                    ]
                }
            ]
        }
    """
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")

    # Build header index map
    header_row = [str(c).strip() if c else "" for c in rows[0]]
    col_map = {}
    for idx, h in enumerate(header_row):
        for expected in EXPECTED_HEADERS:
            if expected in h:
                col_map[expected] = idx
                break

    missing = [h for h in EXPECTED_HEADERS if h not in col_map]
    if missing:
        raise ValueError(f"Missing columns in Excel: {', '.join(missing)}")

    # Group by water plant name, preserving order
    plants: OrderedDict[str, dict] = OrderedDict()

    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue

        wp_name = _cell_str(row, col_map["水厂名称"])
        if not wp_name:
            continue

        wp_scale = _cell_str(row, col_map["水厂规模"])

        if wp_name not in plants:
            plants[wp_name] = {"name": wp_name, "scale": wp_scale, "items": []}

        freq_type = _cell_str(row, col_map["频率类型"])
        freq_value = _cell_int(row, col_map["频率值"], default=1)
        unit_price = _cell_float(row, col_map["单价"])
        annual_count = _cell_int(row, col_map["年检测次数"])
        subtotal = _cell_float(row, col_map["小计"])

        item = {
            "sample_type": _normalize_sample_type(_cell_str(row, col_map["样品类别"])),
            "detection_project": _normalize_detection_project(_cell_str(row, col_map["检测项目"])),
            "detection_standard": _cell_str(row, col_map["检测标准"]),
            "frequency_type": freq_type,
            "frequency_value": freq_value,
            "unit_price": unit_price,
            "annual_count": annual_count,
            "subtotal": subtotal,
        }
        plants[wp_name]["items"].append(item)

    wb.close()
    return {"water_plants": list(plants.values())}


def import_contract_from_excel(
    db: Session,
    company_id: int,
    contract_data: dict,
    file_bytes: bytes,
) -> dict:
    """Create contract, water plants, detection items from parsed Excel,
    then generate annual plan.

    contract_data keys: contract_no, year, start_date, end_date, total_amount
    """
    parsed = parse_excel(file_bytes)

    # Create contract
    contract = Contract(
        contract_no=contract_data["contract_no"],
        company_id=company_id,
        year=contract_data.get("year"),
        start_date=contract_data.get("start_date"),
        end_date=contract_data.get("end_date"),
        total_amount=contract_data.get("total_amount"),
    )
    db.add(contract)
    db.flush()

    total_items = 0
    total_plants = 0

    for wp_data in parsed["water_plants"]:
        wp = WaterPlant(
            name=wp_data["name"],
            scale=wp_data.get("scale"),
            contract_id=contract.id,
        )
        db.add(wp)
        db.flush()
        total_plants += 1

        for item_data in wp_data["items"]:
            di = DetectionItem(
                water_plant_id=wp.id,
                sample_type=item_data.get("sample_type"),
                detection_project=item_data.get("detection_project"),
                detection_standard=item_data.get("detection_standard"),
                frequency_type=item_data.get("frequency_type"),
                frequency_value=item_data.get("frequency_value", 1),
                unit_price=item_data.get("unit_price"),
                annual_count=item_data.get("annual_count"),
                subtotal=item_data.get("subtotal"),
            )
            db.add(di)
            total_items += 1

    db.commit()

    # Generate annual plan
    tasks_count = generate_annual_plan(db, contract.id)

    return {
        "contract_id": contract.id,
        "contract_no": contract.contract_no,
        "water_plants_count": total_plants,
        "detection_items_count": total_items,
        "tasks_generated": tasks_count,
    }


def generate_template() -> bytes:
    """Generate an Excel template with the expected headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "合同费用表"

    for col_idx, header in enumerate(EXPECTED_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Add a sample row
    sample = [
        "示例水厂",
        "10万吨/日",
        "出厂水",
        "月检43项",
        "GB5749常规43项",
        "月",
        1,
        500,
        12,
        0.6,
    ]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---- normalization ----

def _normalize_sample_type(s: str) -> str:
    if not s:
        return s
    if "管网末梢" in s:
        return "管网水"
    if re.search(r'水源水[（(]原水[）)]', s):
        return "水源水"
    return s


def _normalize_detection_project(s: str) -> str:
    if not s:
        return s
    m = re.search(r'(\d+)\s*项', s)
    if m:
        return f"{m.group(1)}项"
    return s


# ---- helpers ----

def _cell_str(row: tuple, idx: int, default: str = "") -> str:
    if idx < len(row) and row[idx] is not None:
        return str(row[idx]).strip()
    return default


def _cell_int(row: tuple, idx: int, default: int = 0) -> int:
    if idx < len(row) and row[idx] is not None:
        try:
            return int(float(row[idx]))
        except (ValueError, TypeError):
            return default
    return default


def _cell_float(row: tuple, idx: int, default: float = 0.0) -> float:
    if idx < len(row) and row[idx] is not None:
        try:
            return float(row[idx])
        except (ValueError, TypeError):
            return default
    return default
